"""
Generate an Excel workbook with sample data and 7 chart types.
Designed for Class 8 exam — large fonts, visible data labels,
clear axis titles, bold legends, values on every data point.
"""
import openpyxl
from openpyxl.chart import (
    BarChart, LineChart, AreaChart, PieChart, DoughnutChart,
    ScatterChart, Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── DATA SHEET ──────────────────────────────────────────────
ws_data = wb.active
ws_data.title = "Data"

# Simple quarterly sales data — easy numbers for Class 8
headers = ["Quarter", "Electronics", "Clothing", "Groceries"]
data = [
    ["Q1", 120, 80, 150],
    ["Q2", 180, 110, 140],
    ["Q3", 150, 130, 170],
    ["Q4", 210, 160, 190],
]

ws_data.append(headers)
for row in data:
    ws_data.append(row)

# Style the data table — large, readable
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
header_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
data_font = Font(size=13, name="Arial")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thick_border = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)
for col in range(1, 5):
    cell = ws_data.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border
    ws_data.column_dimensions[get_column_letter(col)].width = 18
ws_data.row_dimensions[1].height = 30

for row_idx in range(2, 6):
    ws_data.row_dimensions[row_idx].height = 25
    for col in range(1, 5):
        cell = ws_data.cell(row=row_idx, column=col)
        cell.border = thick_border
        cell.font = data_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col >= 2:
            cell.number_format = '#,##0'

# Add a note explaining the data
ws_data["A8"] = "Data: Quarterly sales of 3 products (in ₹ Lakhs)"
ws_data["A8"].font = Font(size=12, italic=True, color="444444", name="Arial")

# References reused across charts
cats = Reference(ws_data, min_col=1, min_row=2, max_row=5)
vals_elec = Reference(ws_data, min_col=2, min_row=1, max_row=5)
vals_cloth = Reference(ws_data, min_col=3, min_row=1, max_row=5)
vals_groc = Reference(ws_data, min_col=4, min_row=1, max_row=5)

# ── HELPER: Big, readable chart styling ─────────────────────
def make_big_font(size=1200, bold=False):
    """Return RichText with the given font size (in 100ths of a point)."""
    rp = CharacterProperties(sz=size, b=bold)
    rp.latin = DrawingFont(typeface="Arial")
    pp = ParagraphProperties(defRPr=rp)
    return RichText(p=[Paragraph(pPr=pp, endParaRPr=rp)])

def style_chart(chart, title, x_title=None, y_title=None, width=22, height=15):
    """Apply exam-friendly styling: large title, axis labels, gridlines, data labels."""
    chart.title = title
    chart.width = width
    chart.height = height

    # Large bold title
    chart.title = title
    title_font = CharacterProperties(sz=1600, b=True)
    title_font.latin = DrawingFont(typeface="Arial")
    chart.title.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(defRPr=title_font),
            endParaRPr=title_font
        )]
    )

    # Legend: large text, at bottom
    chart.legend.position = "b"
    leg_font = CharacterProperties(sz=1100, b=True)
    leg_font.latin = DrawingFont(typeface="Arial")
    chart.legend.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(defRPr=leg_font),
            endParaRPr=leg_font
        )]
    )

    # Axis titles & tick labels (large, readable)
    if x_title and hasattr(chart, 'x_axis'):
        chart.x_axis.title = x_title
        chart.x_axis.title.txPr = make_big_font(1200, bold=True)
        chart.x_axis.txPr = make_big_font(1000)
        chart.x_axis.delete = False
    if y_title and hasattr(chart, 'y_axis'):
        chart.y_axis.title = y_title
        chart.y_axis.title.txPr = make_big_font(1200, bold=True)
        chart.y_axis.txPr = make_big_font(1000)
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines = None  # cleaner look

def add_value_labels(chart):
    """Add data value labels to every series (visible numbers on bars/lines)."""
    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showCatName = False
        s.dLbls.showSerName = False
        lbl_font = CharacterProperties(sz=1000, b=True)
        lbl_font.latin = DrawingFont(typeface="Arial")
        s.dLbls.txPr = RichText(
            p=[Paragraph(
                pPr=ParagraphProperties(defRPr=lbl_font),
                endParaRPr=lbl_font
            )]
        )

# ── 1. BAR CHART (horizontal bars) ─────────────────────────
ws1 = wb.create_sheet("1. Bar Chart")
ws1["B1"] = "BAR CHART — Horizontal bars comparing values across categories"
ws1["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
c1 = BarChart()
c1.type = "bar"
c1.style = 10
c1.add_data(vals_elec, titles_from_data=True)
c1.add_data(vals_cloth, titles_from_data=True)
c1.add_data(vals_groc, titles_from_data=True)
c1.set_categories(cats)
style_chart(c1, "Quarterly Sales (Bar Chart)", x_title="Sales (in ₹ Lakhs)", y_title="Quarter")
add_value_labels(c1)
ws1.add_chart(c1, "B3")

# ── 2. LINE CHART ──────────────────────────────────────────
ws2 = wb.create_sheet("2. Line Chart")
ws2["B1"] = "LINE CHART — Shows trends over time using connected data points"
ws2["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
c2 = LineChart()
c2.style = 10
c2.add_data(vals_elec, titles_from_data=True)
c2.add_data(vals_cloth, titles_from_data=True)
c2.add_data(vals_groc, titles_from_data=True)
c2.set_categories(cats)
style_chart(c2, "Quarterly Sales Trend (Line Chart)", x_title="Quarter", y_title="Sales (in ₹ Lakhs)")
# Thick lines with big markers
for s in c2.series:
    s.graphicalProperties.line.width = 35000  # ~2.5pt
    s.marker.symbol = "circle"
    s.marker.size = 10
add_value_labels(c2)
ws2.add_chart(c2, "B3")

# ── 3. AREA CHART ──────────────────────────────────────────
ws3 = wb.create_sheet("3. Area Chart")
ws3["B1"] = "AREA CHART — Like a line chart but fills the area below the line"
ws3["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
c3 = AreaChart()
c3.style = 10
c3.add_data(vals_elec, titles_from_data=True)
c3.add_data(vals_cloth, titles_from_data=True)
c3.add_data(vals_groc, titles_from_data=True)
c3.set_categories(cats)
style_chart(c3, "Quarterly Sales Volume (Area Chart)", x_title="Quarter", y_title="Sales (in ₹ Lakhs)")
add_value_labels(c3)
ws3.add_chart(c3, "B3")

# ── 4. PIE CHART ───────────────────────────────────────────
ws4 = wb.create_sheet("4. Pie Chart")
ws4["B1"] = "PIE CHART — Shows each part as a percentage of the whole"
ws4["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
# Write summary data on this sheet too
ws4["A3"] = "Product"
ws4["B3"] = "Total Sales"
ws4["A3"].font = Font(bold=True, size=12, name="Arial")
ws4["B3"].font = Font(bold=True, size=12, name="Arial")
ws4["A4"] = "Electronics"
ws4["B4"] = sum(row[1] for row in data)  # 660
ws4["A5"] = "Clothing"
ws4["B5"] = sum(row[2] for row in data)  # 480
ws4["A6"] = "Groceries"
ws4["B6"] = sum(row[3] for row in data)  # 650
for r in range(3, 7):
    for c in range(1, 3):
        cell = ws4.cell(row=r, column=c)
        cell.font = Font(size=12, name="Arial")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thick_border
    ws4.cell(row=r, column=1).font = Font(size=12, bold=(r == 3), name="Arial")
ws4["A3"].fill = header_fill
ws4["A3"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
ws4["B3"].fill = header_fill
ws4["B3"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["B"].width = 14

pie_cats = Reference(ws4, min_col=1, min_row=4, max_row=6)
pie_vals = Reference(ws4, min_col=2, min_row=3, max_row=6)
c4 = PieChart()
c4.add_data(pie_vals, titles_from_data=True)
c4.set_categories(pie_cats)
c4.title = "Total Sales by Product (Pie Chart)"
c4.width = 20
c4.height = 15
c4.legend.position = "b"
# Large title
title_font4 = CharacterProperties(sz=1600, b=True)
title_font4.latin = DrawingFont(typeface="Arial")
c4.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=title_font4), endParaRPr=title_font4)])
# Large legend
leg_font4 = CharacterProperties(sz=1100, b=True)
leg_font4.latin = DrawingFont(typeface="Arial")
c4.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=leg_font4), endParaRPr=leg_font4)])
# Data labels: show category name + value + percentage
c4.dataLabels = DataLabelList()
c4.dataLabels.showPercent = True
c4.dataLabels.showVal = True
c4.dataLabels.showCatName = True
c4.dataLabels.showSerName = False
c4.dataLabels.separator = "\n"
lbl_font4 = CharacterProperties(sz=1100, b=True)
lbl_font4.latin = DrawingFont(typeface="Arial")
c4.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=lbl_font4), endParaRPr=lbl_font4)])
ws4.add_chart(c4, "D3")

# ── 5. DOUGHNUT CHART ──────────────────────────────────────
ws5 = wb.create_sheet("5. Doughnut Chart")
ws5["B1"] = "DOUGHNUT CHART — Like a pie chart but with a hole in the centre"
ws5["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
ws5["A3"] = "Product"
ws5["B3"] = "Total Sales"
ws5["A4"] = "Electronics"
ws5["B4"] = sum(row[1] for row in data)
ws5["A5"] = "Clothing"
ws5["B5"] = sum(row[2] for row in data)
ws5["A6"] = "Groceries"
ws5["B6"] = sum(row[3] for row in data)
for r in range(3, 7):
    for c in range(1, 3):
        cell = ws5.cell(row=r, column=c)
        cell.font = Font(size=12, name="Arial")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thick_border
ws5["A3"].fill = header_fill
ws5["A3"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
ws5["B3"].fill = header_fill
ws5["B3"].font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
ws5.column_dimensions["A"].width = 16
ws5.column_dimensions["B"].width = 14

dn_cats = Reference(ws5, min_col=1, min_row=4, max_row=6)
dn_vals = Reference(ws5, min_col=2, min_row=3, max_row=6)
c5 = DoughnutChart()
c5.add_data(dn_vals, titles_from_data=True)
c5.set_categories(dn_cats)
c5.title = "Total Sales by Product (Doughnut Chart)"
c5.width = 20
c5.height = 15
c5.legend.position = "b"
title_font5 = CharacterProperties(sz=1600, b=True)
title_font5.latin = DrawingFont(typeface="Arial")
c5.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=title_font5), endParaRPr=title_font5)])
leg_font5 = CharacterProperties(sz=1100, b=True)
leg_font5.latin = DrawingFont(typeface="Arial")
c5.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=leg_font5), endParaRPr=leg_font5)])
c5.dataLabels = DataLabelList()
c5.dataLabels.showPercent = True
c5.dataLabels.showVal = True
c5.dataLabels.showCatName = True
c5.dataLabels.showSerName = False
c5.dataLabels.separator = "\n"
lbl_font5 = CharacterProperties(sz=1100, b=True)
lbl_font5.latin = DrawingFont(typeface="Arial")
c5.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=lbl_font5), endParaRPr=lbl_font5)])
ws5.add_chart(c5, "D3")

# ── 6. COLUMN CHART (vertical bars) ────────────────────────
ws6 = wb.create_sheet("6. Column Chart")
ws6["B1"] = "COLUMN CHART — Vertical bars comparing values side by side"
ws6["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
c6 = BarChart()
c6.type = "col"
c6.style = 10
c6.add_data(vals_elec, titles_from_data=True)
c6.add_data(vals_cloth, titles_from_data=True)
c6.add_data(vals_groc, titles_from_data=True)
c6.set_categories(cats)
style_chart(c6, "Quarterly Sales (Column Chart)", x_title="Quarter", y_title="Sales (in ₹ Lakhs)")
add_value_labels(c6)
ws6.add_chart(c6, "B3")

# ── 7. XY SCATTER CHART ────────────────────────────────────
ws7 = wb.create_sheet("7. XY Scatter")
ws7["B1"] = "XY SCATTER CHART — Shows relationship between two sets of numbers"
ws7["B1"].font = Font(size=13, bold=True, color="2F5496", name="Arial")
# Write scatter data on this sheet
ws7["A3"] = "Electronics (X)"
ws7["B3"] = "Clothing (Y)"
ws7["C3"] = "Groceries (Y)"
for col_idx in range(1, 4):
    cell = ws7.cell(row=3, column=col_idx)
    cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thick_border
    ws7.column_dimensions[get_column_letter(col_idx)].width = 18
for i, row in enumerate(data, start=4):
    ws7.cell(row=i, column=1, value=row[1]).font = Font(size=12, name="Arial")
    ws7.cell(row=i, column=2, value=row[2]).font = Font(size=12, name="Arial")
    ws7.cell(row=i, column=3, value=row[3]).font = Font(size=12, name="Arial")
    for c in range(1, 4):
        ws7.cell(row=i, column=c).alignment = Alignment(horizontal="center")
        ws7.cell(row=i, column=c).border = thick_border

c7 = ScatterChart()
c7.style = 10
x_vals = Reference(ws7, min_col=1, min_row=4, max_row=7)
y_vals1 = Reference(ws7, min_col=2, min_row=4, max_row=7)
s1 = Series(y_vals1, x_vals, title="Clothing vs Electronics")
s1.marker.symbol = "circle"
s1.marker.size = 12
c7.series.append(s1)
y_vals2 = Reference(ws7, min_col=3, min_row=4, max_row=7)
s2 = Series(y_vals2, x_vals, title="Groceries vs Electronics")
s2.marker.symbol = "diamond"
s2.marker.size = 12
c7.series.append(s2)
style_chart(c7, "Product Correlation (XY Scatter Chart)",
            x_title="Electronics Sales (₹ Lakhs)", y_title="Other Product Sales (₹ Lakhs)")
# Add value labels to scatter points
for s in c7.series:
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.showCatName = False
    s.dLbls.showSerName = False
    lbl_font_sc = CharacterProperties(sz=1000, b=True)
    lbl_font_sc.latin = DrawingFont(typeface="Arial")
    s.dLbls.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=lbl_font_sc), endParaRPr=lbl_font_sc)]
    )
ws7.add_chart(c7, "E3")

# ── SAVE ────────────────────────────────────────────────────
output = r"C:\Users\SandeepTiwari\MCX\Sample_Charts.xlsx"
wb.save(output)
print(f"✅ Saved: {output}")
